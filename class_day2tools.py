#!/usr/bin/env python3

from datetime import datetime
from easy_functions import variablesFromAPI
from easy_functions import vlan_list_format
from intersight.api import asset_api
from intersight.api import compute_api
from intersight.api import fabric_api
from intersight.api import fcpool_api
from intersight.api import macpool_api
from intersight.api import organization_api
from intersight.api import server_api
from intersight.api import vnic_api
from intersight.exceptions import ApiException
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from pathlib import Path
import credentials
import json
import pkg_resources
import pprint
import pytz
import openpyxl
import re
import sys
import validating
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

pp = pprint.PrettyPrinter(indent=4)
home = Path.home()
template_path = pkg_resources.resource_filename('class_vmware', 'Templates/')

class intersight(object):
    def __init__(self, type):
        self.type = type

    def add_vlan(self, **kwargs):
        args = kwargs['args']
        jsonFile = False
        if not args.json_file == None:
            jsonFile = True
            jsonOpen = open(args.json_file, 'r')
            jsonData = json.load(jsonOpen)
        tags = [{'key': 'Module','value': 'day2tools'}]

        def print_api_class(api_response):
            datastr = str(api_response)
            datastr = datastr.replace('\'', '\"')
            datastr = datastr.replace('None,', 'null,')
            datastr = datastr.replace('False', 'false')
            datastr = datastr.replace('True', 'true')
            ct = re.search('(\"create_time\": )(datetime.datetime\([0-9, ]+tzinfo=tzutc\(\)\))', datastr)
            ctime = '%s"%s",' % (ct.group(1), ct.group(2))
            datastr = re.sub(
                '\"create_time\": datetime.datetime\([0-9, ]+tzinfo=tzutc\(\)\),',
                ctime,
                datastr
            )
            mt = re.search('(\"mod_time\": )(datetime.datetime\([0-9, ]+tzinfo=tzutc\(\)\))', datastr)
            mtime = '%s"%s",' % (mt.group(1), mt.group(2))
            datastr = re.sub(
                '\"mod_time\": datetime.datetime\([0-9, ]+tzinfo=tzutc\(\)\),',
                mtime,
                datastr
            )
            data = json.loads(datastr)
            print(json.dumps(data, indent=4))

        def process_results(apiQuery):
            api_dict = {}
            api_list = []
            empty = False
            if apiQuery.results:
                for i in apiQuery.results:
                    iMoid = i.moid
                    iName = i.name
                    idict = {iName:{'moid':iMoid}}
                    api_dict.update(idict)
                    api_list.append(iName)
                return empty, api_dict, api_list
            else:
                empty = True
                return empty, api_dict, api_list
        
        def empty_results(apiQuery):
                print(f'The API Query Results were empty for {apiQuery.object_type}.  Exiting...')
                exit()

        print(f'\n-------------------------------------------------------------------------------------------\n')
        print('  Beginning VLAN Addition...')
        print(f'\n-------------------------------------------------------------------------------------------\n')

        # Prompt User for VLAN.
        valid = False
        while valid == False:
            vlan_id = '%s' % (input(f'    What is the VLAN ID: '))
            if re.search('^\d+$', vlan_id):
                valid = validating.number_in_range('VLAN ID', vlan_id, 1, 4094)

        # Prompt User for VLAN Name.
        valid = False
        while valid == False:
            vlan_name = '%s' % (input(f'    What is the name you want to assign to VLAN {vlan_id}: '))
            valid = validating.name_rule('VLAN Name', vlan_name, 1, 62)

        # Query API for the Organization List
        api_client = credentials.config_credentials(home, args)
        api_handle = organization_api.OrganizationApi(api_client)
        apiQuery = api_handle.get_organization_organization_list()
        empty, orgs, org_names = process_results(apiQuery)
        if empty == True: empty_results(apiQuery)

        # Request from User Which Organizations to Apply this to.
        if jsonFile == False:
            kwargs["multi_select"] = True
            kwargs["var_description"] = 'Select the Organizations to Apply this VLAN to.'
            kwargs["jsonVars"] = org_names
            kwargs["defaultVar"] = 'default'
            kwargs["varType"] = 'Organizations'
            organizations = variablesFromAPI(**kwargs)
        else:
            organizations = jsonData['organizations']
        for org in organizations:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Starting Loop on Organization {org}.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            # Query the API for the VLAN Policies
            api_handle = fabric_api.FabricApi(api_client)
            query_filter = f"Organization.Moid eq '{orgs[org]['moid']}'"
            qargs = dict(filter=query_filter)
            apiQuery = api_handle.get_fabric_eth_network_policy_list(**qargs)
            empty, vlan_policies,vlan_policies_names = process_results(apiQuery)
            if empty == True: empty_results(apiQuery)

            # Prompt the User to Select the VLAN Policy
            if jsonFile == False:
                kwargs["multi_select"] = False
                kwargs["var_description"] = f'Select the VLAN Policy for Organization {org}.'
                kwargs["jsonVars"] = vlan_policies_names
                kwargs["defaultVar"] = ''
                kwargs["varType"] = 'VLAN Policy'
                vlan_policy = variablesFromAPI(**kwargs)
            else:
                vlan_policy = jsonData['vlan_policy']

            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Checking VLAN Policy {vlan_policy} for VLAN {vlan_id}.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            vlan_policy_moid = vlan_policies[vlan_policy]['moid']
            query_filter = f"EthNetworkPolicy.Moid eq '{vlan_policy_moid}'"
            kwargs = dict(filter=query_filter,top=1000)
            apiQuery = api_handle.get_fabric_vlan_list(**kwargs)
            if apiQuery.results:
                mcast_policy_moid = apiQuery.results[1]['multicast_policy']['moid']
                match_count = 0
                vlan_list = []
                for i in apiQuery.results:
                    vlan_list.append(i['vlan_id'])
                    if int(i['vlan_id']) == int(vlan_id):
                        match_count += 1
                        print(f'\n-------------------------------------------------------------------------------------------\n')
                        print(f'  VLAN is already in the policy {vlan_policy} and has moid {i["moid"]}.')
                        print(f'\n-------------------------------------------------------------------------------------------\n')
                vlan_list.sort()
                vlans = vlan_list_format(vlan_list)
                if match_count == 0:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  VLAN {vlan_id} was not in VLAN Policy {vlan_policy}.  Adding VLAN...')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    policy = {
                        'class_id':'fabric.Vlan',
                        'eth_network_policy': {
                            'class_id': 'mo.MoRef',
                            'moid': vlan_policy_moid,
                            'object_type': 'fabric.EthNetworkPolicy'
                        },
                        'is_native': False,
                        'multicast_policy': {
                            'class_id': 'mo.MoRef',
                            'moid': mcast_policy_moid,
                            'object_type': 'fabric.MulticastPolicy'
                        },
                        'name': vlan_name,
                        'object_type': 'fabric.Vlan',
                        'vlan_id': int(vlan_id)
                        }
                    try:
                        apiPost = api_handle.create_fabric_vlan(policy)
                        print_api_class(apiPost)
                    except ApiException as e:
                        print("Exception when calling FabricApi->create_fabric_vlan: %s\n" % e)
                        sys.exit(1)

            # Query the API for Ethernet Network Group Policies
            query_filter = f"Organization.Moid eq '{orgs[org]['moid']}'"
            qargs = dict(filter=query_filter)
            apiQuery = api_handle.get_fabric_eth_network_group_policy_list(**qargs)
            empty, eth_group_policies,eth_group_policies_names = process_results(apiQuery)
            if empty == True: empty_results(apiQuery)

            # Prompt the User to Select the Ethernet Network Group Policies
            if jsonFile == False:
                kwargs["multi_select"] = True
                kwargs["var_description"] = f'Select the Ethernet Network Group Polices to append'\
                    f'the VLAN to in Organization {org}.'
                kwargs["jsonVars"] = eth_group_policies_names
                kwargs["defaultVar"] = ''
                kwargs["varType"] = 'VLAN Group Policies'
                ethernet_network_group_policies = variablesFromAPI(**kwargs)
            else:
                ethernet_network_group_policies = jsonData['ethernet_network_group_policies']

            for i in ethernet_network_group_policies:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f"  Patching the Ethernet Network Group Policy '{i}' ...")
                print(f'\n-------------------------------------------------------------------------------------------\n')
                ethgroup_moid = eth_group_policies[i]['moid']
                patch_body = {
                    'vlan_settings':{
                        'allowed_vlans':vlans
                    }
                }
                try:
                    apiPatch = api_handle.patch_fabric_eth_network_group_policy(
                        fabric_eth_network_group_policy=patch_body,
                        moid=ethgroup_moid
                    )
                    print_api_class(apiPatch)
                except ApiException as e:
                    print("Exception when calling FabricApi->patch_fabric_eth_network_group_policy: %s\n" % e)
                    sys.exit(1)

            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Checking if the Ethernet Network Group Policy {vlan_id}_NIC-A Already Exists.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            ethgcount = 0
            for k, v in eth_group_policies.items():
                if k == f'{vlan_id}_NIC-A':
                    ethgcount += 1
                    vnic_eth_net_grp = v['moid']
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  Ethernet Network Group Policy {vlan_id}_NIC-A Exists.  Moid is {vnic_eth_net_grp}')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            if ethgcount == 0:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Ethernet Network Group Policy {vlan_id}_NIC-A does not exist.  Creating...')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                policy = {
                    'class_id': 'fabric.EthNetworkGroupPolicy',
                    'description': f'{vlan_id}_NIC-A Ethernet Network Group',
                    'name': f'{vlan_id}_NIC-A',
                    'object_type': 'fabric.EthNetworkGroupPolicy',
                    'organization': {
                        'class_id': 'mo.MoRef',
                        'moid':  orgs[org]['moid'],
                        'object_type': 'organization.Organization'
                    },
                    'tags': tags,
                    'vlan_settings': {
                        'allowed_vlans': f'{vlan_id}',
                        'class_id': 'fabric.VlanSettings',
                        'native_vlan': int(vlan_id),
                        'object_type': 'fabric.VlanSettings'
                    }
                }
                try:
                    apiPost = api_handle.create_fabric_eth_network_group_policy(policy)
                    print_api_class(apiPost)
                    vnic_eth_net_grp = apiPost.moid
                except ApiException as e:
                    print("Exception when calling FabricApi->create_fabric_eth_network_group_policy: %s\n" % e)
                    sys.exit(1)

            # Query the API for the Ethernet Network Control Policies
            query_filter = f"Organization.Moid eq '{orgs[org]['moid']}'"
            qargs = dict(filter=query_filter)
            apiQuery = api_handle.get_fabric_eth_network_control_policy_list(**qargs)
            empty, eth_control_policies,eth_control_names = process_results(apiQuery)
            if empty == True: empty_results(apiQuery)

            # Prompt User for the Ethernet Network Control Policy
            if jsonFile == False:
                kwargs["multi_select"] = False
                kwargs["var_description"] = f'Select the Ethernet Network Control Policy for Organization {org}.'
                kwargs["jsonVars"] = eth_control_names
                kwargs["defaultVar"] = ''
                kwargs["varType"] = 'Ethernet Network Control Policies'
                ethernet_network_control_policy = variablesFromAPI(**kwargs)
            else:
                ethernet_network_control_policy = jsonData['ethernet_network_control_policy']

            # Query the API for MAC Pools
            query_filter = f"Organization.Moid eq '{orgs[org]['moid']}'"
            qargs = dict(filter=query_filter)
            api_handle = macpool_api.MacpoolApi(api_client)
            apiQuery = api_handle.get_macpool_pool_list(**qargs)
            empty, mac_pools,mac_pools_names = process_results(apiQuery)
            if empty == True: empty_results(apiQuery)

            # Prompt User for MAC Pool
            if jsonFile == False:
                kwargs["var_description"] = f'Select the MAC Pool for Organization {org}.'
                kwargs["jsonVars"] = mac_pools_names
                kwargs["defaultVar"] = ''
                kwargs["varType"] = 'MAC Pool'
                mac_pool = variablesFromAPI(**kwargs)
            else:
                mac_pool = jsonData['mac_pool']

            # Query the API for the Ethernet Adapter Policies
            query_filter = f"Organization.Moid eq '{orgs[org]['moid']}'"
            qargs = dict(filter=query_filter)
            api_handle = vnic_api.VnicApi(api_client)
            apiQuery = api_handle.get_vnic_eth_adapter_policy_list(**qargs)
            empty, adapter_policies,adapter_policies_names = process_results(apiQuery)
            if empty == True: empty_results(apiQuery)

            # Prompt User for Ethernet Adapter Policy
            if jsonFile == False:
                kwargs["multi_select"] = False
                kwargs["var_description"] = f'Select the Ethernet Adapter Policy for Organization {org}.'
                kwargs["jsonVars"] = adapter_policies_names
                kwargs["defaultVar"] = ''
                kwargs["varType"] = 'Ethernet Adapter Policies'
                ethernet_adapter_policy = variablesFromAPI(**kwargs)
            else:
                ethernet_adapter_policy = jsonData['ethernet_adapter_policy']


            # Query the API for the Ethernet QoS Policies
            query_filter = f"Organization.Moid eq '{orgs[org]['moid']}'"
            qargs = dict(filter=query_filter)
            apiQuery = api_handle.get_vnic_eth_qos_policy_list(**qargs)
            empty, qos_policies,qos_policies_names = process_results(apiQuery)
            if empty == True: empty_results(apiQuery)

            # Prompt User for Ethernet Adapter Policy
            if jsonFile == False:
                kwargs["var_description"] = f'Select the Ethernet QoS Policy for Organization {org}.'
                kwargs["jsonVars"] = qos_policies_names
                kwargs["defaultVar"] = ''
                kwargs["varType"] = 'Ethernet QoS Policies'
                ethernet_qos_policy = variablesFromAPI(**kwargs)
            else:
                ethernet_qos_policy = jsonData['ethernet_qos_policy']


            # Query the API for the Adapter Policies
            query_filter = f"Organization.Moid eq '{orgs[org]['moid']}'"
            qargs = dict(filter=query_filter)
            apiQuery = api_handle.get_vnic_lan_connectivity_policy_list(**qargs)
            empty, lan_policies,lan_policies_names = process_results(apiQuery)
            if empty == True: empty_results(apiQuery)

            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Checking if the LAN Policy {vlan_id} Already Exists.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            lcount = 0
            for k, v in lan_policies.items():
                if k == vlan_id:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  LAN Policy {vlan_id} exists.  Moid is {v["moid"]}')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    lcount += 1
            if lcount == 0:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  LAN Policy {vlan_id} does not exist.  Creating...')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                policy = {
                    'class_id':'vnic.LanConnectivityPolicy',
                    'name': str(vlan_id),
                    'object_type': 'vnic.LanConnectivityPolicy',
                    'organization': {
                        'class_id': 'mo.MoRef',
                        'moid': orgs[org]['moid'],
                        'object_type': 'organization.Organization'
                    },
                    'tags': tags,
                    'target_platform': 'FIAttached'
                    }
                try:
                    apiPost = api_handle.create_vnic_lan_connectivity_policy(policy)
                    print_api_class(apiPost)
                    lanp = {vlan_id:{'moid':apiPost.moid}}
                    lan_policies.update(lanp)
                except ApiException as e:
                    print("Exception when calling VnicApi->create_vnic_lan_connectivity_policy: %s\n" % e)
                    sys.exit(1)

            print(f'\n-------------------------------------------------------------------------------------------\n')
            print('  Checking if the LAN Connectivity vNIC "NIC-A" Exists...')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            query_filter = f"LanConnectivityPolicy.Moid eq '{lan_policies[vlan_id]['moid']}'"
            qargs = dict(filter=query_filter)
            apiQuery = api_handle.get_vnic_eth_if_list(**qargs)
            empty, vnics, vnic_names = process_results(apiQuery)
            if empty == True:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f"  vNIC 'NIC-A' was not attached to LAN Policy {vlan_id}.  Creating...")
                print(f'\n-------------------------------------------------------------------------------------------\n')
                policy = {
                    'cdn': {
                        'class_id': 'vnic.Cdn',
                        'object_type': 'vnic.Cdn',
                        'source': 'vnic',
                        'value': 'NIC-A'
                    },
                    'class_id': 'vnic.EthIf',
                    'eth_adapter_policy': {
                        'class_id': 'mo.MoRef',
                        'moid': adapter_policies[ethernet_adapter_policy]['moid'],
                        'object_type': 'vnic.EthAdapterPolicy'
                    },
                    'eth_network_policy': None,
                    'eth_qos_policy': {
                        'class_id': 'mo.MoRef',
                        'moid': qos_policies[ethernet_qos_policy]['moid'],
                        'object_type': 'vnic.EthQosPolicy'
                    },
                    'fabric_eth_network_control_policy': {
                        'class_id': 'mo.MoRef',
                        'moid': eth_control_policies[ethernet_network_control_policy]['moid'],
                        'object_type': 'fabric.EthNetworkControlPolicy'
                    },
                    'fabric_eth_network_group_policy': [
                        {
                            'class_id': 'mo.MoRef',
                            'moid': vnic_eth_net_grp,
                            'object_type': 'fabric.EthNetworkGroupPolicy'
                        }
                    ],
                    'failover_enabled': True,
                    'lan_connectivity_policy': {
                        'class_id': 'mo.MoRef',
                        'moid': lan_policies[vlan_id]['moid'],
                        'object_type': 'vnic.LanConnectivityPolicy'
                    },
                    'mac_address_type': 'POOL',
                    'mac_lease': None,
                    'mac_pool': {
                        'class_id': 'mo.MoRef',
                        'moid': mac_pools[mac_pool]['moid'],
                        'object_type': 'macpool.Pool'
                    },
                    'name': 'NIC-A',
                    'object_type': 'vnic.EthIf',
                    'order': 3,
                    'placement': {
                        'class_id': 'vnic.PlacementSettings',
                        'id': 'MLOM',
                        'object_type': 'vnic.PlacementSettings',
                        'pci_link': 0,
                        'switch_id': 'A',
                        'uplink': 0
                    },
                }
                try:
                    apiPost = api_handle.create_vnic_eth_if(policy)
                    print_api_class(apiPost)
                except ApiException as e:
                    print("Exception when calling VnicApi->create_vnic_lan_connectivity_policy: %s\n" % e)
                    sys.exit(1)
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  LAN Connectivity vNIC "NIC-A" exists.  Moid is {vnics["NIC-A"]["moid"]}')
                print(f'\n-------------------------------------------------------------------------------------------\n')
                
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Finished Loop on Organization {org}.')
            print(f'\n-------------------------------------------------------------------------------------------\n')

        print(f'\n-------------------------------------------------------------------------------------------\n')
        print('  Finished Adding VLAN...')
        print(f'\n-------------------------------------------------------------------------------------------\n')

    def server_inventory(self, **kwargs):
        args = kwargs['args']
        pyDict = {}
        # Obtain Server Profile Data
        api_client = credentials.config_credentials(home, args)
        kwargs = dict(top=1000)
        api_handle = server_api.ServerApi(api_client)
        apiQuery = api_handle.get_server_profile_list(**kwargs)
        if apiQuery.results:
            apiResults = apiQuery.results
            for i in apiResults:
                profileMoid = i.moid
                profileName = i.name
                print(f'Obtaining Data for Server Profile {profileName}')
                if i.target_platform == 'FIAttached':
                    if i.associated_server:
                        api_handle = fcpool_api.FcpoolApi(api_client)
                        query_filter = f"PoolPurpose eq 'WWNN' and AssignedToEntity.Moid eq '{profileMoid}'"
                        kwargs = dict(filter=query_filter)
                        fcPool = api_handle.get_fcpool_lease_list(**kwargs)
                        wwnn_address = fcPool.results[0].wwn_id

                        # Obtain Physical UCS Server Information
                        serverMoid = i.associated_server.moid
                        serverType = i.associated_server.object_type
                        api_handle = compute_api.ComputeApi(api_client)
                        if serverType == 'compute.Blade':
                            apiQuery = api_handle.get_compute_blade_by_moid(serverMoid)
                            serverDn = 'chassis-' + str(apiQuery.chassis_id) + "/blade-" + str(apiQuery.slot_id)
                        else:
                            apiQuery = api_handle.get_compute_rack_unit_by_moid(serverMoid)
                            serverDn = 'rackunit-' + str(apiQuery.server_id)
                        serverSerial = apiQuery.serial
                        api_handle = asset_api.AssetApi(api_client)
                        serverReg = api_handle.get_asset_device_registration_by_moid(apiQuery.registered_device.moid)
                        domainParent = api_handle.get_asset_device_registration_by_moid(serverReg.parent_connection.moid)
                        # Obtain Server vNICs (Ethernet/Fibre-Channel)
                        api_handle = vnic_api.VnicApi(api_client)
                        query_filter = f"Profile.Moid eq '{profileMoid}'"
                        kwargs = dict(filter=query_filter)
                        eth_apiQuery = api_handle.get_vnic_eth_if_list(**kwargs)
                        fc_apiQuery = api_handle.get_vnic_fc_if_list(**kwargs)
                        vnics = {}
                        for item in eth_apiQuery.results:
                            vnic_name = item['name']
                            mac_address = item['mac_address']
                            ngpMoid = item['fabric_eth_network_group_policy'][0].moid
                            qosMoid = item['eth_qos_policy']['moid']
                            api_handle = vnic_api.VnicApi(api_client)
                            qosPolicy = api_handle.get_vnic_eth_qos_policy_by_moid(qosMoid)
                            mTu = qosPolicy.mtu
                            vnic = {
                                vnic_name: {
                                    'mac_address':mac_address,
                                    'mtu':mTu,
                                }
                            }
                            vnics.update(vnic)

                        vhbas = {'A':{},'B':{}}
                        for item in fc_apiQuery.results:
                            vhba_name = item['name']
                            if item['wwpn_address_type'] == 'STATIC':
                                wwpn_address = item['static_wwpn_address']
                            else:
                                wwpn_address = item['wwpn']
                            switch_id = item['placement']['switch_id']
                            fcnpMoid = item['fc_network_policy'].moid
                            vhbas[switch_id] = {
                                'vhba':vhba_name,
                                'wwpn_address':wwpn_address
                            }
                        hostResults = {
                            profileName:{
                                'domain_name':domainParent.device_hostname[0],
                                'moid':profileMoid,
                                'serial':serverSerial,
                                'server_dn':serverDn,
                                'vhbas':vhbas,
                                'vnics':vnics,
                                'wwnn':wwnn_address
                            }
                        }
                        pyDict.update(hostResults)

            # print(json.dumps(pyDict, indent=4))
            # Build Named Style Sheets for Workbook
            bd1 = Side(style="thick", color="0070C0")
            bd2 = Side(style="medium", color="0070C0")
            heading_1 = NamedStyle(name="heading_1")
            heading_1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
            heading_1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
            heading_1.fill = PatternFill("solid", fgColor="305496")
            heading_1.font = Font(bold=True, size=15, color="FFFFFF")
            heading_2 = NamedStyle(name="heading_2")
            heading_2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
            heading_2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
            heading_2.font = Font(bold=True, size=15, color="44546A")
            even = NamedStyle(name="even")
            even.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
            even.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
            even.font = Font(bold=False, size=12, color="44546A")
            odd = NamedStyle(name="odd")
            odd.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
            odd.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
            odd.fill = PatternFill("solid", fgColor="D9E1F2")
            odd.font = Font(bold=False, size=12, color="44546A")

            Est = pytz.timezone('US/Eastern')
            datetime_est = datetime.now(Est)
            Est1 = datetime_est.strftime('%Y-%m-%d_%H-%M')
            Est2 = datetime_est.strftime('%Y-%m-%d %H:%M:%S %Z %z')
            workbook = f'UCS_WWPN_Collector-{Est1}.xlsx'
            wb = openpyxl.Workbook()
            wb.add_named_style(heading_1)
            wb.add_named_style(heading_2)
            wb.add_named_style(even)
            wb.add_named_style(odd)
            ws = wb.active
            ws.title = 'WWPN List'
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws_header = f'Collected UCS Data on {Est2}'
            data = [ws_header]
            ws.append(data)
            ws.merge_cells('A1:D1')
            for cell in ws['1:1']:
                cell.style = 'heading_1'
            data = [
                'Server Profile','WWNN','WWPN for Fabric A','WWPN for Fabric B'
            ]
            ws.append(data)
            for cell in ws['2:2']:
                cell.style = 'heading_2'
            ws_row_count = 3
            for k, v in pyDict.items():
                if v['wwnn']: wwnn = v['wwnn']
                else: wwnn = 'Not Configured'
                if v['vhbas']['A'].get('wwpn_address'):
                    vhba1_wwpn = v['vhbas']['A']['wwpn_address']
                else:
                    vhba1_wwpn = 'Not Configured'
                if v['vhbas']['B'].get('wwpn_address'):
                    vhba2_wwpn = v['vhbas']['B']['wwpn_address']
                else:
                    vhba2_wwpn = 'Not Configured'
                data = [
                    k,
                    wwnn,
                    vhba1_wwpn,
                    vhba2_wwpn
                ]
                ws.append(data)
                for cell in ws[ws_row_count:ws_row_count]:
                    if ws_row_count % 2 == 0:
                        cell.style = 'odd'
                    else:
                        cell.style = 'even'
                ws_row_count += 1
            wb.save(filename=workbook)
